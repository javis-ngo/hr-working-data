from copy import copy
from openpyxl.cell import WriteOnlyCell
from openpyxl.formula.translate import Translator
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook import Workbook
from functools import lru_cache
import os


TEMPLATE_FILE = os.path.abspath("workday_data.xlsx")
wb_template = load_workbook(TEMPLATE_FILE, data_only=False, keep_vba=False)

@lru_cache(maxsize=8)
def parse_template(template_path: str, sheet_name: str, header_skip_rows: int, keep_formula_columns: tuple[str, ...]):
    ws = wb_template[sheet_name]
    header_row_idx = header_skip_rows + 1
    data_start_row = header_row_idx + 1

    header_row = ws[header_row_idx]
    header_map = {}
    header_styles = []
    formula_templates = {}

    for col_idx, cell in enumerate(header_row, 1):
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        header_map[header] = col_idx
        # COPY STYLE OBJECTS to avoid StyleProxy hashing issues
        header_styles.append({
            'font': copy(cell.font) if cell.font else Font(),
            'fill': copy(cell.fill) if cell.fill else PatternFill(),
            'border': copy(cell.border) if cell.border else Border(),
            'alignment': copy(cell.alignment) if cell.alignment else Alignment(),
        })
        if header in keep_formula_columns:
            tcell = ws.cell(row=data_start_row, column=col_idx)
            if isinstance(tcell.value, str) and tcell.value.startswith("="):
                formula_templates[header] = tcell.value

    # Cache các dòng pre-header (value + style refs)
    pre_header_rows = []
    for row_idx in range(1, header_row_idx):
        row_cells = []
        for col_idx in range(1, len(header_row) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            row_cells.append({
                'value': c.value,
                'font': copy(c.font) if c.font else Font(),
                'fill': copy(c.fill) if c.fill else PatternFill(),
                'border': copy(c.border) if c.border else Border(),
                'alignment': copy(c.alignment) if c.alignment else Alignment(),
            })
        pre_header_rows.append(row_cells)

    ordered_headers = list(header_map.keys())
    return {
        'header_map': header_map,
        'ordered_headers': ordered_headers,
        'header_styles': header_styles,
        'formula_templates': formula_templates,
        'pre_header_rows': pre_header_rows,
        'header_row_len': len(header_row),
        'data_start_row': data_start_row,
    }

def write_preserving_formulas_and_styles(template_path, output_path, df, sheet_name, header_skip_rows, keep_formula_columns):
    tpl = parse_template(template_path, sheet_name, header_skip_rows, tuple(keep_formula_columns))
    header_map = tpl['header_map']
    ordered_headers = tpl['ordered_headers']
    header_styles = tpl['header_styles']
    pre_header_rows = tpl['pre_header_rows']
    formula_templates = tpl['formula_templates']
    data_start_row = tpl['data_start_row']

    num_rows = len(df)
    df_values = {h: (df[h].values if h in df.columns else [None]*num_rows) for h in ordered_headers}

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    # append pre-header rows (value + style refs)
    for row_meta in pre_header_rows:
        row = []
        for meta in row_meta:
            nc = WriteOnlyCell(ws, value=meta['value'])
            nc.font = meta['font']; nc.fill = meta['fill']; nc.border = meta['border']; nc.alignment = meta['alignment']
            row.append(nc)
        ws.append(row)

    # header row
    header_cells = []
    for header, style in zip(ordered_headers, header_styles):
        c = WriteOnlyCell(ws, value=header)
        c.font = style['font']; c.fill = style['fill']; c.border = style['border']; c.alignment = style['alignment']
        header_cells.append(c)
    ws.append(header_cells)

    # precompute translated formulas per column (prefer fast path)
    translated_formulas = {}
    for header, formula in formula_templates.items():
        col_idx = header_map[header]
        col_letter = get_column_letter(col_idx)
        base_ref = f"{col_letter}{data_start_row}"
        # fast path (same-col refs)
        if base_ref in formula:
            translated_formulas[header] = [formula.replace(base_ref, f"{col_letter}{data_start_row + i}") for i in range(num_rows)]
        else:
            # fallback Translator once per column
            tr = Translator(formula, origin=f"{col_letter}{data_start_row}")
            translated_formulas[header] = [tr.translate_formula(dest=f"{col_letter}{data_start_row + i}") for i in range(num_rows)]

    # write data rows
    for ridx in range(num_rows):
        row = []
        for header in ordered_headers:
            cell = WriteOnlyCell(ws)
            if header in translated_formulas:
                cell.value = translated_formulas[header][ridx]
            else:
                cell.value = df_values[header][ridx]
            row.append(cell)
        ws.append(row)

    wb.save(output_path)
    wb.close()